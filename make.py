import sys
import os
import tempfile
import base64
import shutil
from pathlib import Path
from openpyxl import load_workbook, Workbook

# Optional: Validate with pyxform if available
try:
    from pyxform.xls2xform import xls2xform_convert
    from pyxform.errors import PyXFormError
    PYXFORM_AVAILABLE = True
except ImportError:
    PYXFORM_AVAILABLE = False

def validate_with_pyxform(xls_path):
    print("\n🔍 Validating XLSForm using pyxform...")
    java_home = os.environ.get("JAVA_HOME")
    if java_home:
        java_bin = os.path.join(java_home, "bin")
        if java_bin not in os.environ["PATH"]:
            os.environ["PATH"] = java_bin + os.pathsep + os.environ["PATH"]

    try:
        with tempfile.NamedTemporaryFile(suffix=".xml", delete=False) as tmp:
            tmp_path = tmp.name
        # Note: older versions don't accept kwargs like output_path or validate=True
        xls2xform_convert(str(xls_path), tmp_path)
        print("✅ XLSForm is valid!")
    except PyXFormError as e:
        print("❌ Validation failed:")
        print(e)
    except Exception as e:
        print("❌ Unexpected error during validation:")
        print(e)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

# Copy media files
def copy_media_files(form_path: Path, build_dir: Path):
    """
    Copy all files from the 'media' subdirectory next to form_path into build_dir.
    Only flat copy — subdirectories are ignored.
    """
    media_src = form_path.parent / "media"
    if not media_src.exists() or not media_src.is_dir():
        print("📭 No media directory found.")
        return

    print(f"📁 Copying media files from {media_src} to {build_dir}")
    for file in media_src.iterdir():
        if file.is_file():
            try:
                shutil.copy(file, build_dir / file.name)
            except Exception as e:
                print(f"⚠️ Failed to copy {file.name}: {e}")

# Inject the qml files
def process_xlsform(input_path):
    input_path = Path(input_path).resolve()
    form_dir = input_path.parent
    output_dir = form_dir / "build"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / input_path.name

    print(f"📄 Loading: {input_path}")
    workbook = load_workbook(input_path)
    if "survey" not in workbook.sheetnames:
        print("❌ No 'survey' sheet found.")
        return

    survey = workbook["survey"]

    # Identify header row and column indices
    headers = {cell.value: i for i, cell in enumerate(next(survey.iter_rows(min_row=1, max_row=1)))}
    if "bind::ct:content.qmlFile" not in headers:
        print("❌ Column 'bind::ct:content.qmlFile' not found.")
        return

    col_old = headers["bind::ct:content.qmlFile"]
    col_new = max(headers.values()) + 1
    survey.cell(row=1, column=col_new + 1).value = "bind::ct:content.qmlBase64"

    # Process rows
    for row in survey.iter_rows(min_row=2):
        cell = row[col_old]
        if cell.value:
            qml_file = form_dir / cell.value
            try:
                content = qml_file.read_text(encoding="utf-8")
                content = base64.b64encode(content.encode('utf-8')).decode('ascii')
                row[col_new].value = content
            except Exception as e:
                print(f"⚠️ Failed to read {qml_file}: {e}")
                row[col_new].value = f"# ERROR: could not read {cell.value}"

    # Build new workbook, excluding the old column
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "survey"

    for r_idx, row in enumerate(survey.iter_rows(values_only=True), 1):
        new_row = [val for i, val in enumerate(row) if i != col_old]
        new_ws.append(new_row)

    # Copy all other sheets unchanged
    for sheet_name in workbook.sheetnames:
        if sheet_name == "survey":
            continue
        src = workbook[sheet_name]
        dst = new_wb.create_sheet(title=sheet_name)
        for row in src.iter_rows(values_only=True):
            dst.append(row)

    # Save to output
    new_wb.save(output_path)
    print(f"✅ Saved processed XLSForm to: {output_path}")

    # Run validation
    if PYXFORM_AVAILABLE:
        validate_with_pyxform(output_path)
    else:
        print("\n⚠️ pyxform is not installed; skipping validation.")
        print("To enable validation, run: pip install pyxform")

    # Copy media files
    copy_media_files(input_path, output_dir)

# Main
if __name__ == "__main__":
    if len(sys.argv) == 1:
        form_file = "form.xlsx"
        print("📂 No input file specified, defaulting to 'form.xlsx'")
    elif len(sys.argv) == 2:
        form_file = sys.argv[1]
    else:
        print("Usage: python process_xlsform.py [path/to/form.xlsx]")
        sys.exit(1)
        
    process_xlsform(form_file)
